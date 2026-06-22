from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import requests
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

app = Flask(__name__)
CORS(app)

EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'contacts.xlsx')
GOOGLE_APPS_SCRIPT_URL = os.getenv('GOOGLE_APPS_SCRIPT_URL')


def init_excel():
    """Initialize Excel file with headers if it doesn't exist."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        headers = ['Date', 'Full Name', 'Email', 'Phone', 'Company', 'Message']
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = cell.font.copy(bold=True)
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 50
        wb.save(EXCEL_FILE)


@app.route('/api/contact', methods=['POST'])
def submit_contact():
    """Handle contact form submission and save to Google Sheets (with Excel fallback)."""
    try:
        data = request.get_json()

        # Validate required fields
        required_fields = ['name', 'email', 'phone', 'message']
        for field in required_fields:
            if not data.get(field, '').strip():
                return jsonify({
                    'success': False,
                    'message': f'Field "{field}" is required.'
                }), 400

        saved_to_sheets = False
        
        # Try sending to Google Sheets if URL is configured
        if GOOGLE_APPS_SCRIPT_URL and GOOGLE_APPS_SCRIPT_URL.strip():
            try:
                payload = {
                    'name': data.get('name', '').strip(),
                    'email': data.get('email', '').strip(),
                    'phone': data.get('phone', '').strip(),
                    'company': data.get('company', '').strip(),
                    'message': data.get('message', '').strip()
                }
                
                # Make POST request with a timeout
                response = requests.post(GOOGLE_APPS_SCRIPT_URL, json=payload, timeout=10)
                
                if response.status_code == 200:
                    resp_json = response.json()
                    if resp_json.get('success'):
                        saved_to_sheets = True
                    else:
                        print(f"Google Sheets Script error: {resp_json.get('error')}")
                else:
                    print(f"Google Sheets Script returned status code {response.status_code}")
            except Exception as sheet_err:
                print(f"Failed to save to Google Sheets: {sheet_err}")

        # If saving to Google Sheets failed or URL is not configured, save to local Excel
        if not saved_to_sheets:
            print("Google Sheets integration not active or failed. Saving to local Excel fallback.")
            init_excel()
            
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

            row_data = [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                data.get('name', '').strip(),
                data.get('email', '').strip(),
                data.get('phone', '').strip(),
                data.get('company', '').strip(),
                data.get('message', '').strip()
            ]

            ws.append(row_data)
            wb.save(EXCEL_FILE)
            
            if GOOGLE_APPS_SCRIPT_URL and GOOGLE_APPS_SCRIPT_URL.strip():
                msg = 'Saved to local backup (Google Sheets sync failed).'
            else:
                msg = 'Saved locally. Configure GOOGLE_APPS_SCRIPT_URL to sync to Google Sheets.'
        else:
            msg = 'Contact information saved to Google Sheets successfully!'

        return jsonify({
            'success': True,
            'message': msg
        }), 200

    except Exception as e:
        print(f"Error saving contact: {e}")
        return jsonify({
            'success': False,
            'message': 'An error occurred while saving your information.'
        }), 500


@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint."""
    return jsonify({'status': 'ok', 'message': 'MH TraCon API is running'})


if __name__ == '__main__':
    init_excel()
    print("MH TraCon Backend running on http://localhost:5000")
    app.run(debug=True, port=5000)

